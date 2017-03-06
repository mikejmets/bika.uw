# -*- coding: utf-8 -*-
import os
from Products.ATContentTypes.utils import dt2DT

from bika.lims.browser.arimport.handler import ImportHandler as BaseHandler
from bika.lims.content.analysisservice import AnalysisService
from bika.lims.interfaces import IARImportHandler
from bika.lims.utils import tmpID
from bika.lims.utils.analysisrequest import create_analysisrequest

from DateTime import DateTime
from Products.Archetypes.event import ObjectInitializedEvent
from Products.CMFCore.utils import getToolByName
from Products.CMFPlone.utils import _createObjectByType
from tempfile import mkstemp
from zExceptions import Redirect
from zope import event
from zope.interface import implements

import openpyxl
import transaction


class ImportHandler(BaseHandler):
    """Override arimport behaviour
    """
    implements(IARImportHandler)

    def __init__(self, context, request):
        super(ImportHandler, self).__init__(context, request)

    def parse_raw_data(self):
        """Parse the UW-format of import file, do some limited parsing,
        and store the values

        This is intentionally brittle, the UW form requires certain
        fields at certain positions, and is fixed.  Further modifications
        to the input file require modifications to this code.
        """

        context = self.context
        bika_catalog = getToolByName(context, 'bika_catalog')
        bika_setup_catalog = getToolByName(context, 'bika_setup_catalog')

        blob = context.Schema()['RawData'].get(context)

        wb, fn = self.load_workbook(blob.data)
        ws = wb.worksheets[0]

        # Header data
        _clientname = ws['C2'].value
        _clientid = ws['D2'].value
        _contactname = ws['E2'].value
        _clientordernumber = ws['F2'].value
        _clientreference = ws['G2'].value
        # Batch data
        _batchtitle = ws['B4'].value
        _batchid = ws['C4'].value
        _batchdescription = ws['D4'].value
        _clientbatchid = ws['E4'].value
        returnsampletoclient = ws['F4'].value
        # "Batch meta" (just more batch data really)
        _clientbatchcomment = ws['B6'].value
        _labbatchcomment = ws['C6'].value
        # analytes (profile titles, service titles, service keywords, CAS nr)
        _analytes = [ws[chr(x) + '8'].value
                     for x in range(66, 91)  # B=66, Z=90
                     if ws[chr(x) + '8'].value]
        # Samples "meta" (common values for all samples)
        try:
            _datesampled = str(dt2DT(ws['B10'].value))
        except:
            _datesampled = ''
        _sampletype = ws['C10'].value
        _samplepoint = ws['D10'].value
        _activitysampled = ws['E10'].value
        # count the number of sample rows
        nr_samples = 0
        while 1:
            if ws['B{}'.format(12 + nr_samples)].value:
                nr_samples += 1
            else:
                break

        # If batch already exists, link it now.
        brains = bika_catalog(portal_type='Batch', title=_batchtitle)
        batch = brains[0].getObject() if brains else None

        # Lookup sample type and point
        sampletype = None
        sampletypes = bika_setup_catalog(
            portal_type='SampleType',
            title=_sampletype)
        if len(sampletypes) == 1:
            sampletype = sampletypes[0].getObject()

        samplepoint = None
        samplepoints = bika_setup_catalog(
            portal_type='SamplePoint',
            title=_samplepoint)
        if len(samplepoints) == 1:
            samplepoint = samplepoints[0].getObject()

        # Write applicable values to ARImport schema
        # These are values that will be used in all created objects,
        # and are set only once.
        arimport_values = {
            'ClientName': _clientname,
            'ClientID': _clientid,
            'ClientOrderNumber': _clientordernumber,
            'ClientReference': _clientreference,
            'ContactName': _contactname,
            'CCContacts': [],
            'SamplePoint': samplepoint,
            'SampleType': sampletype,
            'ActivitySampled': _activitysampled,
            'BatchTitle': _batchtitle,
            'BatchDescription': _batchdescription,
            'BatchID': _batchid,
            'ClientBatchID': _clientbatchid,
            'LabBatchComment': _labbatchcomment,
            'ClientBatchComment': _clientbatchcomment,
            'Batch': batch,
            'NrSamples': nr_samples,
        }
        # Write initial values to ARImport schema
        for fieldname, fieldvalue in arimport_values.items():
            context.Schema()[fieldname].set(context, fieldvalue)

        itemdata = []
        for sample_nr in range(nr_samples):
            clientsampleid = ws['B{}'.format(12 + sample_nr)].value
            amountsampled = ws['C{}'.format(12 + sample_nr)].value
            metric = ws['D{}'.format(12 + sample_nr)].value
            remarks = ws['E{}'.format(12 + sample_nr)].value
            values = {
                'ClientSampleID': str(clientsampleid),
                'AmountSampled': str(amountsampled),
                'AmountSampledMetric': str(metric),
                'DateSampled': _datesampled,
                'Analyses': _analytes,
                'Remarks': remarks,
                # 'Profile': self.profile,
            }
            itemdata.append(values)
            context.Schema()['ItemData'].set(context, itemdata)
        context.reindexObject()

        # Close worksheet and remove the tmp file.
        wb.close()
        os.unlink(fn)

    def load_workbook(self, data):
        fd, fn = mkstemp(suffix=".xlsx")
        os.close(fd)
        open(fn, 'wb').write(data)
        wb = openpyxl.load_workbook(fn, data_only=True, guess_types=False)
        return wb, fn

    def validate(self):
        """Resolve and validate stored values

        This function is responsible for setting context.Valid=True.

        If this function does not set Valid=True, the reasons why must
        be stored in context.Errors.
        """

        context = self.context
        pc = getToolByName(context, 'portal_catalog')
        bc = getToolByName(context, 'bika_catalog')

        errors = []

        # Validate Client info
        client = None
        clientname = context.Schema()['ClientName'].get(context)
        clientid = context.Schema()['ClientID'].get(context)
        brains = pc(portal_type='Client', getName=clientname)
        if brains:
            # Client name found: validate client's ID against import file
            client = brains[0].getObject()
            if clientid and clientid != client.getClientID():
                errors.append(
                    "Client '{}' has ID '{}', but you specified '{}'".format(
                        client.Title(), client.getClientID(), clientid))
        else:
            # Client name not found
            errors.append(
                "Client name is invalid; please select a valid client, "
                "or enter a valid Client Name.")
        if client:
            if context.aq_parent != client:
                # Wrong client name specified
                errors.append("Selected client '{}' should be '{}'!".format(
                    client.Title(), context.aq_parent.Title()))

        # Validate ContactName
        if client:
            contactname = context.Schema()['ContactName'].get(context)
            if not self.is_valid_contact(contactname):
                errors.append(
                    "Contact name is not a valid contact for this client.")
        else:
            errors.append(
                "Cannot validate Contact until valid Client is selected.")

        # Validate integrity of the different batch fields if the
        # Batch ID or Title reference an existing batch.
        batch = None

        batch_id = context.getBatchID()
        batch_title = context.getBatchTitle()
        # First try to find existing batch
        brains = bc(portal_type='Batch', id=batch_id)
        if not brains:
            brains = bc(portal_type='Batch', title=batch_title)
        if brains:
            batch = brains[0].getObject()
            # if both title and id are specified, make sure they both match
            if batch_title and batch_id:
                if batch.Title() != batch_title:
                    errors.append("Existing Work Order title is not {}".format(
                        batch_title
                    ))
                if batch.getId() != batch_id:
                    errors.append("Existing Work Order ID is not {}".format(
                        batch_id
                    ))

        # Simple SamplePoint validation
        sp = context.Schema()['SamplePoint'].get(context)
        if isinstance(sp, basestring):
            errors.append("'{}' is not a valid sample point.".format(sp))
        if not sp:
            errors.append(
                "The selected sample point/sampling location was not found.")

        # Simple SamplePoint validation
        st = context.Schema()['SampleType'].get(context)
        if isinstance(st, basestring):
            errors.append("'{}' is not a valid sample point.".format(st))
        if not st:
            errors.append(
                "The selected sample medium/sample type was not found.")

        # Validate ItemData fields
        for item in context.Schema()['ItemData'].get(self.context):
            csid = item['ClientSampleID']
            # Analytes
            for analyte in item['Analyses']:
                uids = self.resolve_analyses(analyte)
                if isinstance(uids, basestring):
                    errors.append("{}: {}".format(csid, uids))
            # DateSampled
            try:
                DateTime(item['DateSampled'])
            except:
                errors.append("{}: Invalid date {}".format(
                    csid, item['DateSampled']))

        context.Schema()['Errors'].set(context, errors)
        if not errors:
            context.Schema()['Valid'].set(context, True)

        valid = context.Schema()['Valid'].get(context)
        if not valid:
            self.statusmessage("There were errors during validation.")
            url = self.context.absolute_url() + "/base_edit"
            transaction.commit()
            raise Redirect(url)

    def is_valid_contact(self, fullname):
        contacts = self.context.aq_parent.objectValues('Contact')
        for contact in contacts:
            if contact.getFullname() == fullname:
                return True
        return False

    def import_items(self):

        context = self.context
        request = context.REQUEST
        uc = getToolByName(context, 'uid_catalog')
        bika_catalog = getToolByName(context, 'bika_catalog')

        client = context.aq_parent
        contact = [c for c in client.objectValues('Contact')
                   if c.getFullname() == self.context.getContactName()][0]

        self.progressbar_init('Submitting AR Import')

        # Find existing batch or create new batch if required.
        batch = None
        batch_id = context.getBatchID()
        batch_title = context.getBatchTitle()
        # First try to find existing batch
        brains = bika_catalog(portal_type='Batch', id=batch_id)
        if not brains:
            brains = bika_catalog(portal_type='Batch', title=batch_title)
        if brains:
            batch = brains[0]
        if not batch:
            # Create batch if it does not exist
            _bid = batch_id if batch_id else tmpID()
            batch = _createObjectByType("Batch", client, _bid)
            batch.unmarkCreationFlag()
            batch.edit(
                title=batch_title,
                description=context.getBatchDescription(),
                ClientBatchID=context.getClientBatchID(),
                Remarks=context.getLabBatchComment(),
                ClientBatchComment=context.getClientBatchComment()
            )
            if not batch_id:
                batch._renameAfterCreation()
            event.notify(ObjectInitializedEvent(batch))
            batch.at_post_create_script()

        itemdata = context.Schema()['ItemData'].get(context)

        for i, item in enumerate(itemdata):
            service_uids = []
            for a in item['Analyses']:
                for service in self.resolve_analyses(a):
                    if isinstance(service, AnalysisService):
                        service_uids.append(service.UID())

            # Create AR
            ar_values = {
                'Contact': contact,
                'ClientOrderNumber': context.getClientOrderNumber(),
                'Remarks': item['Remarks'],
                'Batch': batch if batch else None,
                'ClientReference': context.getClientReference(),
                'ClientSampleID': item['ClientSampleID'],
                'SampleType': context.getSampleType(),
                'SamplePoint': context.getSamplePoint(),
                'DateSampled': DateTime(item['DateSampled']),
                'SamplingDate': DateTime(item['DateSampled']),
                'Remarks': item['Remarks'],
            }
            ar = create_analysisrequest(
                client, request, ar_values, analyses=service_uids,
                partitions=[{}]
            )

            self.progressbar_progress(i + 1, len(itemdata))
